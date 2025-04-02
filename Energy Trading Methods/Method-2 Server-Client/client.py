import socket
import sys
import threading
import time
import json
import uuid
import queue
import math
from datetime import datetime, timedelta
# Import openpyxl for Excel logging (if available)
try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None

# Try to import RPi.GPIO; if not available (e.g., on a PC), simulate the functions.
try:
    import RPi.GPIO as GPIO
except ImportError:
    class GPIO:
        BCM = None
        OUT = None
        HIGH = 1
        LOW = 0
        @staticmethod
        def setmode(mode):
            pass
        @staticmethod
        def setup(pin, mode):
            pass
        @staticmethod
        def output(pin, state):
            pass
        @staticmethod
        def cleanup():
            pass

class P2PClient:
    def __init__(self, server_ip, server_port):
        self.server_ip = server_ip
        self.server_port = server_port
        self.username = None
        self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        self.response_queue = queue.Queue()  # For synchronous responses

    def connect_to_server(self):
        try:
            self.sock.connect((self.server_ip, self.server_port))
            print(f"Connected to server at {self.server_ip}:{self.server_port}")
            # Start a background thread to continuously receive messages.
            self.receiver_thread = threading.Thread(target=self.receive_messages, daemon=True)
            self.receiver_thread.start()
        except Exception as e:
            print(f"Error connecting to server: {e}")
            sys.exit(1)

    def receive_messages(self):
        """
        Continuously receive messages from the server.
        If a message is a transaction notification, handle it immediately.
        Otherwise, place the message into the response queue for synchronous requests.
        """
        while True:
            try:
                data = self.sock.recv(4096)
                if not data:
                    # Connection closed
                    break
                try:
                    msg = json.loads(data.decode())
                except Exception as e:
                    print("Error decoding message:", e)
                    continue

                # If the message is a transaction notification, handle it immediately.
                if msg.get("type") == "TRANSACTION_NOTIFICATION":
                    duration = msg.get("duration")
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                    if duration is not None:
                        hours, remainder = divmod(duration, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        print(f"\nTransaction completed for {hours} Hours: {minutes} Minutes: {seconds} Seconds at {current_time}.")

                    else:
                        print(f"\nTransaction completed at {current_time}.")
                    if hasattr(self, 'role'):
                        if self.role == 'seller':
                            threading.Thread(target=self.seller_led, args=(duration,), daemon=True).start()
                        elif self.role == 'buyer':
                            threading.Thread(target=self.buyer_led, args=(duration,), daemon=True).start()
                        else:
                            print("Unknown role. No LED triggered.")
                    else:
                        print("Role not set. No LED triggered.")
                else:
                    # Otherwise, place the message in the response queue.
                    self.response_queue.put(msg)
            except socket.error as e:
                print("Socket error in receive_messages:", e)
                break
            except Exception as e:
                print("Error receiving data:", e)
                break

    def get_response(self):
        """
        Block until a response (non-notification) is available in the queue.
        """
        return self.response_queue.get()

    def authenticate(self):
        self.username = input("Enter your username: ")
        password = input("Enter your password: ")
        auth_message = {
            'type': 'AUTH',
            'username': self.username,
            'password': password
        }
        self.sock.sendall(json.dumps(auth_message).encode())
        response_data = self.get_response()
        if response_data.get("status") == "AUTH_SUCCESS":
            print("Authentication successful.")
            return True
        else:
            print("Authentication failed.")
            return False
    '''def seller_led(self, duration):
        """Glows the LED continuously for the specified duration (in seconds)."""
        LED_PIN = 18  # Adjust if needed
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(LED_PIN, GPIO.OUT)
        #print(f"Glowing LED for {duration} seconds to indicate energy received.")
        for i in range((duration)):
            GPIO.output(LED_PIN, GPIO.HIGH)
            time.sleep(1)
            GPIO.output(LED_PIN, GPIO.LOW)
            time.sleep(1)

        GPIO.cleanup()'''
    
    
    
    def buyer_led(self, duration):
        """Glows the LED continuously for the specified duration (in seconds)."""
        LED_PIN = 18  # Adjust if needed
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(LED_PIN, GPIO.OUT)
        print(f"Glowing LED for {duration} seconds to indicate energy received.")
        glow=(int(duration))/2
        for i in range(math.floor(glow)):
            GPIO.output(LED_PIN, GPIO.HIGH)
            time.sleep(1)
            GPIO.output(LED_PIN, GPIO.LOW)
            time.sleep(1)
        GPIO.cleanup()
        
    def seller_led(self, duration):
        """Glows the LED continuously for the specified duration (in seconds)."""
        LED_PIN = 18  # Adjust if needed
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(LED_PIN, GPIO.OUT)
        print(f"Glowing LED for {duration} seconds to indicate energy sending.")
        glow=(int(duration))/2
        
        for i in range(math.floor(glow)):
            GPIO.output(LED_PIN, GPIO.HIGH)
            time.sleep(1)
            GPIO.output(LED_PIN, GPIO.HIGH)
            time.sleep(1)
        GPIO.cleanup()

    
    
    
    
    
    
    '''def buyer_led(self, duration):
        """Glows the LED continuously for the specified duration (in seconds)."""
        LED_PIN = 18  # Adjust if needed
        GPIO.setmode(GPIO.BCM)
        GPIO.setup(LED_PIN, GPIO.OUT)
        print(f"Glowing LED for {duration} seconds to indicate energy received.")
        for i in range((duration)):
            GPIO.output(LED_PIN, GPIO.HIGH)
            time.sleep(1)
            GPIO.output(LED_PIN, GPIO.LOW)
            time.sleep(1)
        # GPIO.output(LED_PIN, GPIO.HIGH)
        # time.sleep(1)
        # GPIO.output(LED_PIN, GPIO.LOW)
        # GPIO.output(LED_PIN, GPIO.HIGH)
        # time.sleep(1)
        # GPIO.output(LED_PIN, GPIO.LOW)
        # GPIO.output(LED_PIN, GPIO.HIGH)
        # time.sleep(1)
        # GPIO.output(LED_PIN, GPIO.LOW)
        # GPIO.output(LED_PIN, GPIO.HIGH)
        # time.sleep(1)
        # GPIO.output(LED_PIN, GPIO.LOW)
        # GPIO.output(LED_PIN, GPIO.HIGH)
        # time.sleep(1)
        # GPIO.output(LED_PIN, GPIO.LOW)
        GPIO.cleanup()'''

    def log_transaction(self, transaction_type, details=""):
        """Log transaction details to an Excel file."""
        if Workbook is None or load_workbook is None:
            print("openpyxl not available. Skipping Excel logging.")
            return
        file_name = "transactions.xlsx"
        try:
            wb = load_workbook(file_name)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            # Write header row.
            ws.append(["Timestamp", "Username", "Transaction_Type", "Details"])
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        ws.append([timestamp, self.username, transaction_type, details])
        wb.save(file_name)
        print("Transaction logged to Excel.")

    def auto_mode(self):
        print("\nSelect Auto Mode type:")
        print("1. Auto Mode as Seller")
        print("2. Auto Mode as Buyer")
        choice = input("Enter your choice (1 or 2): ")
        if choice == '1':
            self.auto_mode_seller()
        elif choice == '2':
            self.auto_mode_buyer()
        else:
            print("Invalid selection for auto mode.")

    def auto_mode_seller(self):
        seller_name = self.username
        self.role = 'seller'
        try:
            energy_amount = float(input("Enter amount of energy available (in kWh): "))
            min_price = float(input("Enter minimum acceptable price per kWh: "))
            duration_input = input("Enter time duration for energy availability (in HH:MM:SS): ")
            duration = self.cal_total_sec(duration_input)
            # Instead of a full transaction date, get the trading window.
            start_time, end_time = self.get_time_window()
        except ValueError:
            print("Invalid numeric input. Auto mode registration aborted.")
            return

        auto_message = {
            'type': 'AUTO_SELLER',
            'seller_id': str(uuid.uuid4()),
            'seller_name': seller_name,
            'energy_amount': energy_amount,
            'min_price': min_price,
            'duration': duration,
            # Use the time window instead of an absolute transaction date.
            'time_window': {'start': start_time, 'end': end_time},
            'username': self.username
        }
        self.sock.sendall(json.dumps(auto_message).encode())
        response_data = self.get_response()
        print("Server response:", response_data.get("status"))

    def auto_mode_buyer(self):
        buyer_name = self.username
        self.role = 'buyer'
        try:
            needed_energy = float(input("Enter the required energy (in kWh): "))
            max_price = float(input("Enter maximum acceptable price per kWh: "))
            duration_input = input("Enter duration (in HH:MM:SS) for which energy should be active: ")
            duration = self.cal_total_sec(duration_input)
            # Instead of asking for a transaction date, get a trading time window.
            start_time, end_time = self.get_time_window()
        except ValueError:
            print("Invalid numeric input. Auto mode transaction aborted.")
            return

        auto_message = {
            'type': 'AUTO_BUYER',
            'buyer_name': buyer_name,
            'needed_energy': needed_energy,
            'max_price': max_price,
            'duration': duration,
            # Include the time window.
            'time_window': {'start': start_time, 'end': end_time},
            'username': self.username
        }
        self.sock.sendall(json.dumps(auto_message).encode())
        response_data = self.get_response()
        print("Server response:", response_data.get("status"))

    def user_interaction_loop(self):
        while True:
            print("\nSelect your role:")
            print("1. Seller")
            print("2. Buyer")
            print("3. Auto Mode Transaction")
            print("4. Exit")
            choice = input("Enter your choice (1, 2, or 3): ")
            if choice == '1':
                self.register_seller()
            elif choice == '2':
                self.buyer_menu()
            elif choice == '4':
                print("Exiting. Thank you for using P2P Energy Trading.")
                try:
                    self.sock.shutdown(socket.SHUT_RDWR)
                except Exception as e:
                    print("Error during socket shutdown:", e)
                self.sock.close()
                sys.exit()

            elif choice == '3':
                self.auto_mode()
            else:
                print("Invalid choice. Please try again.")
    def cal_total_sec(self,duration):
        duration = duration.split(':')
        hours = int(duration[0])
        minutes = int(duration[1])
        seconds = int(duration[2])
        return hours*3600 + minutes*60 + seconds
    
    def register_seller(self):
        # Use the authenticated username as the seller name.
        seller_name = self.username
        energy_type = input("Enter type of energy (e.g., solar, wind): ")
        try:
            energy_amount = float(input("Enter amount of energy available (in kWh): "))
            # Instead of getting a full transaction date, ask for a trading time window.
            start_time, end_time = self.get_time_window()  # New function that returns start and end times (as strings)
            duration_input = input("Enter time duration for energy availability (in HH:MM:SS): ")
            duration = self.cal_total_sec(duration_input)
            price = float(input("Enter price per kWh: "))
        except ValueError:
            print("Invalid numeric input. Registration aborted.")
            return
        # Generate a unique seller ID.
        seller_id = str(uuid.uuid4())
        seller_message = {
            'type': 'SELLER_REGISTER',
            'seller_id': seller_id,
            'seller_name': seller_name,
            'energy_type': energy_type,
            'energy_amount': energy_amount,
            # Instead of 'transaction_date', send a time window dictionary:
            'time_window': {'start': start_time, 'end': end_time},
            'duration': duration,
            'price': price,
            'username': self.username
        }
        self.sock.sendall(json.dumps(seller_message).encode())
        response_data = self.get_response()
        if response_data.get("status") == "seller_registered":
            print(f"You are registered as a seller with unique ID: {seller_id}")
            self.role = 'seller'
            self.seller_menu(seller_id)
        else:
            print("Seller registration failed.")


    def seller_menu(self, seller_id):
        while True:
            #print("\nSeller Menu:")
            print("Please wait for the Buyer to buy energy from you.Else you can choose from following:\n Seller Menu:")
            print("1. Update Price")
            print("2. Update Energy Amount")
            # Notifications are handled automatically by the receiver thread.
            print("3. Exit Seller Menu")
            choice = input("Enter your choice: ")
            if choice == '1':
                try:
                    new_price = float(input("Enter new price per kWh: "))
                except ValueError:
                    print("Invalid input. Please enter a number.")
                    continue
                update_message = {
                    'type': 'SELLER_UPDATE',
                    'seller_id': seller_id,
                    'field': 'price',
                    'value': new_price,
                    'username': self.username
                }
                self.sock.sendall(json.dumps(update_message).encode())
                response_data = self.get_response()
                print(response_data.get("message", "Price updated."))
            elif choice == '2':
                try:
                    new_amount = float(input("Enter new amount of energy available (in kWh): "))
                except ValueError:
                    print("Invalid input. Please enter a number.")
                    continue
                update_message = {
                    'type': 'SELLER_UPDATE',
                    'seller_id': seller_id,
                    'field': 'energy_amount',
                    'value': new_amount,
                    'username': self.username
                }
                self.sock.sendall(json.dumps(update_message).encode())
                response_data = self.get_response()
                print(response_data.get("message", "Energy amount updated."))
            elif choice == '3':
                exit_message = {
                    'type': 'SELLER_EXIT',
                    'seller_id': seller_id,
                    'username': self.username
                }
                self.sock.sendall(json.dumps(exit_message).encode())
                response_data = self.get_response()
                print(response_data.get("message", "Exited seller menu."))
                self.user_interaction_loop()
            else:
                print("Invalid choice. Please try again.")
    '''def get_valid_transaction_date(self):
        while True:
            date_input = input("Enter desired transaction start time (YYYY-MM-DD HH:MM, at least 12 hours ahead): ")
            try:
                transaction_date = datetime.strptime(date_input, "%Y-%m-%d %H:%M")
            except ValueError:
                print("Invalid date format. Please use YYYY-MM-DD HH:MM.")
                continue
            if transaction_date < datetime.now() + timedelta(hours=12):
                print("The transaction date must be at least 12 hours in the future.")
                continue
            return transaction_date'''

    def get_time_window(self):
        """
        Ask the user to input a time window for trading (e.g., "from" and "to" in HH:MM format).
        For a day-ahead market, the assumption is that the window is on the next day.
        """
        while True:
            start_input = input("Enter trading start time (HH:MM, day-ahead): ")
            end_input = input("Enter trading end time (HH:MM, day-ahead): ")
            try:
                start_time = datetime.strptime(start_input, "%H:%M").time()
                end_time = datetime.strptime(end_input, "%H:%M").time()
                if start_time >= end_time:
                    print("Start time must be earlier than end time. Please try again.")
                    continue
                return start_time.isoformat(), end_time.isoformat()
            except ValueError:
                print("Invalid time format. Please use HH:MM.")

    def buyer_menu(self):
        buyer_name = self.username
        self.role = 'buyer'
        try:
            needed_energy = float(input("Enter the required energy (in kWh): "))
            duration = input("Enter duration (in HH:MM:SS) for which energy should be active: ")
            duration_sec = self.cal_total_sec(duration)
        except ValueError:
            print("Invalid numeric input. Transaction aborted.")
            return
        
        # Ask for a desired trading time window.
        start_time, end_time = self.get_time_window()
        # Save these values for later use.
        self.needed_energy = needed_energy
        self.duration = duration_sec
        self.trading_window = {'start': start_time, 'end': end_time}

        buyer_message = {
            'type': 'BUYER_REQUEST',
            'buyer_name': buyer_name,
            'needed_energy': needed_energy,
            'duration': duration_sec,
            'time_window': self.trading_window,
            'username': self.username
        }
        self.sock.sendall(json.dumps(buyer_message).encode())
        response_data = self.get_response()
        available_sellers = response_data.get("available_sellers", [])
        if not available_sellers:
            print("No sellers available with matching time windows.")
            self.buyer_countinued()  # Option to update trading window.
            return

        # List sellers (they should have matching time window info)
        print("\nAvailable Sellers:")
        for idx, seller in enumerate(available_sellers, start=1):
            print(f"{idx}. ID: {seller['seller_id']}, Name: {seller['seller_name']}, "
                f"Time Window: {seller.get('time_window', {})}, Price: {seller['price']} per kWh")
        try:
            choice = int(input("Select a seller by number (0 to cancel): "))
        except ValueError:
            print("Invalid input.")
            return
        if choice == 0:
            print("Transaction canceled.")
            self.buyer_countinued()
            return
        if choice < 1 or choice > len(available_sellers):
            print("Invalid seller selection.")
            return
        selected_seller = available_sellers[choice - 1]
        transaction_message = {
            'type': 'TRANSACTION',
            'buyer_name': buyer_name,
            'seller_id': selected_seller['seller_id'],
            'energy_amount': needed_energy,
            'duration': duration_sec,
            'price': selected_seller['price'],
            'time_window': self.trading_window,
            'username': self.username
        }
        self.sock.sendall(json.dumps(transaction_message).encode())
        response_data = self.get_response()
        if response_data.get("status") == "transaction_success":
            current_time = time.strftime("%Y-%m-%d %H:%M:%S")
            print(f"Transaction completed with seller {selected_seller['seller_name']} at {current_time}.")
            details = (f"Bought {needed_energy} kWh for {duration} from seller "
                    f"{selected_seller['seller_id']} at {selected_seller['price']} per kWh "
                    f"within time window {self.trading_window}.")
            self.log_transaction("Buyer Transaction", details)
        else:
            print("Transaction failed.")
            self.buyer_countinued()

    def buyer_countinued(self):
        while True:
            print("\nDo you want to:")
            print("1. Wait for an updated seller list")
            print("2. Change the amount of energy needed and/or the time window")
            print("3. Exit")
            buyer_choice = input("Enter your choice: ")
            if buyer_choice == '1':
                print("Waiting... Press Enter when you want to refresh the seller list")
                input()
                # Resend the buyer request with stored values.
                buyer_message = {
                    'type': 'BUYER_REQUEST',
                    'buyer_name': self.username,
                    'needed_energy': self.needed_energy,
                    'duration': self.duration,
                    'time_window': self.trading_window,
                    'username': self.username
                }
                self.sock.sendall(json.dumps(buyer_message).encode())
                response_data = self.get_response()
                sellers = response_data.get("available_sellers", [])
                if not sellers:
                    print("No sellers available with matching time window.")
                    continue
                print("\nAvailable Sellers (Updated List):")
                for idx, seller in enumerate(sellers, start=1):
                    print(f"{idx}. ID: {seller['seller_id']}, Name: {seller['seller_name']}, "
                        f"Time Window: {seller.get('time_window', {})}, Price: {seller['price']} per kWh")
                try:
                    choice = int(input("Select a seller by number (0 to cancel): "))
                except ValueError:
                    print("Invalid input. Returning to buyer menu.")
                    continue
                if choice == 0:
                    continue
                if choice < 1 or choice > len(sellers):
                    print("Invalid seller selection.")
                    continue
                selected_seller = sellers[choice - 1]
                transaction_message = {
                    'type': 'TRANSACTION',
                    'buyer_name': self.username,
                    'seller_id': selected_seller['seller_id'],
                    'energy_amount': self.needed_energy,
                    'duration': self.duration,
                    'price': selected_seller['price'],
                    'time_window': self.trading_window,
                    'username': self.username
                }
                self.sock.sendall(json.dumps(transaction_message).encode())
                response_data = self.get_response()
                if response_data.get("status") == "transaction_success":
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                    print(f"Transaction completed with seller {selected_seller['seller_name']} at {current_time}.")
                    details = (f"Bought {self.needed_energy} kWh from seller {selected_seller['seller_id']} "
                            f"at {selected_seller['price']} per kWh within time window {self.trading_window}.")
                    self.log_transaction("Buyer Transaction", details)
                    break
                else:
                    print("Transaction failed. Returning to buyer menu.")
            elif buyer_choice == '2':
                # Allow the buyer to update energy amount and/or time window.
                try:
                    new_amount = float(input("Enter new amount of energy needed (in kWh): "))
                    self.needed_energy = new_amount
                except ValueError:
                    print("Invalid input for energy amount.")
                    continue
                print("Update the trading time window:")
                self.trading_window = {}
                self.trading_window['start'], self.trading_window['end'] = self.get_time_window()
                buyer_message = {
                    'type': 'BUYER_REQUEST',
                    'buyer_name': self.username,
                    'needed_energy': self.needed_energy,
                    'duration': self.duration,
                    'time_window': self.trading_window,
                    'username': self.username
                }
                self.sock.sendall(json.dumps(buyer_message).encode())
                print("Buyer request updated. Please wait for updated seller list.")
                response_data = self.get_response()
                sellers = response_data.get("available_sellers", [])
                if not sellers:
                    print("No sellers available with matching time window.")
                    continue
                print("\nAvailable Sellers (Updated List):")
                for idx, seller in enumerate(sellers, start=1):
                    print(f"{idx}. ID: {seller['seller_id']}, Name: {seller['seller_name']}, "
                        f"Time Window: {seller.get('time_window', {})}, Price: {seller['price']} per kWh")
                try:
                    choice = int(input("Select a seller by number (0 to cancel): "))
                except ValueError:
                    print("Invalid input. Returning to buyer menu.")
                    continue
                if choice == 0:
                    continue
                if choice < 1 or choice > len(sellers):
                    print("Invalid seller selection.")
                    continue
                selected_seller = sellers[choice - 1]
                transaction_message = {
                    'type': 'TRANSACTION',
                    'buyer_name': self.username,
                    'seller_id': selected_seller['seller_id'],
                    'energy_amount': self.needed_energy,
                    'duration': self.duration,
                    'price': selected_seller['price'],
                    'time_window': self.trading_window,
                    'username': self.username
                }
                self.sock.sendall(json.dumps(transaction_message).encode())
                response_data = self.get_response()
                if response_data.get("status") == "transaction_success":
                    current_time = time.strftime("%Y-%m-%d %H:%M:%S")
                    print(f"Transaction completed with seller {selected_seller['seller_name']} at {current_time}.")
                    details = (f"Bought {self.needed_energy} kWh from seller {selected_seller['seller_id']} "
                            f"at {selected_seller['price']} per kWh within time window {self.trading_window}.")
                    self.log_transaction("Buyer Transaction", details)
                    break
                else:
                    print("Transaction failed. Returning to buyer menu.")
            elif buyer_choice == '3':
                print("Exiting buyer menu.")
                break
            else:
                print("Invalid selection, please try again.")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python client.py <server_ip> <server_port>")
        sys.exit(1)
    server_ip = sys.argv[1]
    try:
        server_port = int(sys.argv[2])
    except ValueError:
        print("Invalid port number.")
        sys.exit(1)
    client = P2PClient(server_ip, server_port)
    client.connect_to_server()
    if client.authenticate():
        client.user_interaction_loop()






